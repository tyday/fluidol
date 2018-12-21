import openpyxl, os, requests, urllib
from fluidol_index.models import Product, Category, Product_applications, Product_benefits, Product_features
from fluidol.settings import MEDIA_ROOT
from django.core.management.base import BaseCommand 

# getting urls and turning them into image files
# https://stackoverflow.com/questions/13137817/how-to-download-image-using-requests
# https://stackoverflow.com/questions/1308386/programmatically-saving-image-to-django-imagefield

# import fluidol_index.models.Product
# from ..fluidol_index.models import Product
# from .. import fluidol_index

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

workbook = openpyxl.load_workbook('C:\\Users\\tyrda\\Programming\\fluidol\\fluidol_django\\utilities\\seal_information.xlsx')
amalgamatedSheet = workbook['Amalgamated']
def get_image(image_url):
    name = image_url[(image_url.rfind('/')+1):]
    # response = requests.get(image_url)
    # with open(name,'wb') as f:
    #     f.write(response.content)
    # return f
    result = urllib.request.urlretrieve(image_url)
    # result = urllib.urlretrieve(image_url)
    return result

def save_and_report_image(image_url):
    """ given an URL will save the image in the 
    MEDIA_ROOT and return the name of the image
    """
    name = image_url[(image_url.rfind('/')+1):]
    image_path = MEDIA_ROOT + '\\' + name
    response = requests.get(image_url)
    with open(image_path,'wb') as f:
        f.write(response.content)
    print('Save and report')
    print(image_path + '\n' + name)
    return name

def upload_seal():
    for i in range(2,5): #range(2,51)
        name = amalgamatedSheet.cell(row=i,column=1).value
        if name:
            try:
                category = amalgamatedSheet.cell(row=i,column=2).value.split(',')
            except:
                category = []
            description = amalgamatedSheet.cell(row=i,column=3).value
            try:
                features = amalgamatedSheet.cell(row=i,column=4).value.split(',')
            except:
                features = []
            try:
                benefits = amalgamatedSheet.cell(row=i,column=5).value.split(',')
            except:
                benefits = []
            try:
                applications = amalgamatedSheet.cell(row=i,column=6).value.split(',')
            except:
                applications = []
            try:
                industries = amalgamatedSheet.cell(row=i,column=7).value.split(',')
            except:
                industries = []
            try:
                mainImageTxt = amalgamatedSheet.cell(row=i,column=8).value
                mainImage = save_and_report_image(amalgamatedSheet.cell(row=i,column=9).value)
            except:
                mainImageTxt = None
                mainImage = None
            try:
                optImg1Txt = amalgamatedSheet.cell(row=i,column=10).value
                optImg1 = save_and_report_image(amalgamatedSheet.cell(row=i,column=11).value)
            except:
                optImg1Txt = None
                optImg1 = None
            try:
                optImg2Txt = amalgamatedSheet.cell(row=i,column=12).value
                optImg2 = save_and_report_image(amalgamatedSheet.cell(row=i,column=13).value)
            except:
                optImg2Txt = None
                optImg2 = None
            
            product = Product.objects.create(name=name)
            # product.name = name
            for cat in category:
                # product.category = category
                cat = cat.strip()
                try:
                    findCat = Category.objects.get(name=cat)
                    product.category.add(findCat)
                except:
                    findCat = Category.objects.create(name=cat)
                    product.category.add(findCat)
                # product.category_set.add(cat)
            product.description = description
            for feature in features:
                # product.Product_features = feature
                feature = feature.strip()
                try:
                    findFeat = Product_features.objects.get(name=feature)
                    product.Product_features.add(findFeat)
                except:
                    findFeat = Product_features.objects.create(name=feature)
                    product.Product_features.add(findFeat)
            for benefit in benefits:
            #     product.Product_benefits = benefit
                benefit = benefit.strip()
                try:
                    findBenny = Product_benefits.objects.get(name=benefit)
                    product.Product_benefits.add(findBenny)
                except:
                    findBenny = Product_benefits.objects.create(name=benefit)
                    product.Product_benefits.add(findBenny)
            for application in applications:
                # product.Product_applications = application
                application = application.strip()
                try:
                    findApplication = Product_applications.objects.get(name=application)
                    product.Product_applications.add(findApplication)
                except:
                    findApplication = Product_applications.objects.create(name=application)
                    product.Product_applications.add(findApplication)
                
            # Does not look like industries is in model yet
            # for industry in industries:
            #     product.Product_industries = industry
            product.main_photo_description = mainImageTxt
            product.main_photo = mainImage
            product.optional_photo_one_description = optImg1Txt
            product.optional_photo_one = optImg1
            product.optional_photo_two_description = optImg2Txt
            product.optional_photo_two = optImg2
            #not implemented
            # product.optional_photo_three_description = optImg3Txt
            # product.optional_photo_three = optImg3
            product.save()
            
            print(name,category,len(category))
class Command(BaseCommand):
    def handle(self, **options):
        upload_seal()
if __name__ == '__main__':
    # upload_seal()
    import os
    print(os.path)