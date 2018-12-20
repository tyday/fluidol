import openpyxl

workbook = openpyxl.load_workbook('seal_information.xlsx')
amalgamatedSheet = workbook['Amalgamated']

for i in range(2,51):
    name = amalgamatedSheet.cell(row=i,column=1).value
    if name:
        try:
            category = amalgamatedSheet.cell(row=i,column=2).value.split(',')
        except:
            category = []
        description = amalgamatedSheet.cell(row=i,column=3).value
        try:
            features = amalgamatedSheet.cell(row=i,column=4).value.split(',')
        except:
            features = []
        try:
            benefits = amalgamatedSheet.cell(row=i,column=5).value.split(',')
        except:
            benefits = []
        try:
            applications = amalgamatedSheet.cell(row=i,column=6).value.split(',')
        except:
            applications = []
        try:
            industries = amalgamatedSheet.cell(row=i,column=7).value.split(',')
        except:
            industries = []
        try:
            mainImageTxt = amalgamatedSheet.cell(row=i,column=8).value
            mainImage = amalgamatedSheet.cell(row=i,column=9).value
        except:
            mainImageTxt = None
            mainImage = None
        try:
            optImg1Txt = amalgamatedSheet.cell(row=i,column=8).value
            optImg1 = amalgamatedSheet.cell(row=i,column=9).value
        except:
            optImg1Txt = None
            optImg1 = None
        try:
            optImg2Txt = amalgamatedSheet.cell(row=i,column=8).value
            optImg2 = amalgamatedSheet.cell(row=i,column=9).value
        except:
            optImg2Txt = None
            optImg2 = None
        
        
        
        print(name,category,len(category))